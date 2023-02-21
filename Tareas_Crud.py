from datetime import datetime
from openpyxl import load_workbook

Rut="C:\\Users\\SENA\\Desktop\\git_taller\\crud_python\\tareascrud.xlsx"
Rut = r"C:\Users\SENA\Desktop\git_taller\crud_python\tareascrud.xlsx"
#direcciones de los archivos para poder ejecutar el codigo, importante

def leer(ruta:str, extraer:str):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos_de_crud']
    Hoja_datos = Hoja_datos['A2':'F' + str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:
        if isinstance(i[0].value,int):
            info.setdefault(i[0].value,{'tarea':i[1].value,  'descripcion':i[2].value,
                                        'estado':i[3].value, 'fecha de inicio':i[4].value,
                                        'fecha de finalizacion':i[5].value })
    
    if not(extraer == 'todo'):
        info = filtrar(info, extraer)
    
    for i in info:
        print('********Tarea********')
        print('Id: ' + str(i) + '\n' + 'Titulo:  ' + str(info[i]['tarea']) + '\n' + 'Descripcion: ' + str(info[i]['descripcion']) + '\n'
              'Estado: ' + str(info[i]['estado']) + '\n' + 'Fecha Creacion: ' + str(info[i]['fecha de inicio']) + '\n' + 'Fecha de finalizacion: '+  str(info[i]['fecha de finalizacion']) )    
        print()

    return


def filtrar(info:dict,filtro:str):
    aux={}

    for i in info:
        if info[i]['estado'] == filtro:
            aux.setdefault(i,info[i])

    return aux
    


def actualizar(ruta:str, identificador:int, datos_actualizados:dict):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos_de_crud']
    Hoja_datos = Hoja_datos['A2':'F' + str(Hoja_datos.max_row)]
    hoja = Archivo_Excel.active

    titulo = 2
    descripcion = 3
    estado = 4
    fecha_inicio = 5
    fecha_FInalizado = 6
    encontro = False
    for i in Hoja_datos:
        if i[0].value == identificador:
            fila=i[0].row   
            encontro = True
            for d in datos_actualizados:
                if d == 'titulo' and not (datosactualizados[d] == ''):
                    hoja.cell(row = fila, column = titulo).value= datosactualizados[d]
                elif d == 'descripcion' and not (datosactualizados[d] == ''):
                    hoja.cell(row = fila, column = descripcion).value= datosactualizados[d]
                elif d == 'estado' and not (datosactualizados[d] == ''):
                    hoja.cell(row = fila, column = estado).value= datosactualizados[d]
                elif d == 'fecha inicio' and not (datosactualizados[d] == ''):
                    hoja.cell(row = fila, column = fecha_inicio).value= datosactualizados[d]
                elif d == 'fecha finalizacion' and not (datosactualizados[d] == ''):
                    hoja.cell(row = fila, column = fecha_FInalizado).value= datosactualizados[d]

    Archivo_Excel.save(ruta)
    if encontro == False:
        print('Error: No existe una tarea con ese ID')
        print()
    return

def agregrar(ruta:int, datos:dict):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos_de_crud']
    Hoja_datos = Hoja_datos['A2':'F' + str(Hoja_datos.max_row+1)]
    hoja = Archivo_Excel.active

    titulo = 2
    descripcion = 3
    estado = 4
    fecha_inicio = 5
    fecha_FInalizado = 6
    for i in Hoja_datos:

        if not( isinstance(i[0].value, int )) :
            identificador=i[0].row
            hoja.cell(row = identificador, column =1).value = identificador -1
            hoja.cell(row = identificador, column = titulo).value = datos['titulo']
            hoja.cell(row = identificador, column = descripcion).value = datos['descripcion']
            hoja.cell(row = identificador, column = estado).value = datos['estado']
            hoja.cell(row = identificador, column = fecha_inicio).value = datos['fecha inicio']
            hoja.cell(row = identificador, column = fecha_FInalizado).value = datos['fecha finalizacion']
            break
    Archivo_Excel.save(ruta)
    return

def borrar(ruta, identificador):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['Datos_de_crud']
    Hoja_datos = Hoja_datos['A2':'F' + str(Hoja_datos.max_row)]
    hoja = Archivo_Excel.active

    titulo = 2
    descripcion = 3
    estado = 4
    fecha_inicio = 5
    fecha_FInalizado = 6
    encontro = False
    for i in Hoja_datos:
        if i[0].value == identificador:
            fila = i[0].row
            encontro = True

            hoja.cell(row=fila, column = 1).value = ""
            hoja.cell(row=fila, column = titulo).value = ""
            hoja.cell(row=fila, column = descripcion).value = ""
            hoja.cell(row=fila, column = estado).value = ""
            hoja.cell(row=fila, column = fecha_inicio).value = ""
            hoja.cell(row=fila, column = fecha_FInalizado).value = ""
    Archivo_Excel.save(ruta)
    if encontro == False:
        print("Error: No exite una tarea con ese ID")
        print()
    return

Rut="C:\\Users\\SENA\\Desktop\\git_taller\\crud_python\\tareascrud.xlsx"
#direcciones de los archivos para poder ejecutar el codigo, importante

datosactualizados={'titulo':'','descripcion':'','estado':'','fecha inicio':'','fecha finalizacion':''}
while True:
    print('Indique la acción que desea realizar: ')
    print('Consultar: 1')
    print('Actualizar: 2')
    print('Crear nueva tarea:  3')
    print('Borrar:  4')
    accion = input('escriba la opción: ')

    if not(accion == '1') and not (accion == '2') and not (accion == '3') and not (accion == '4'):
        print('Comando invalido, por favor elija una opcion valida')
    elif accion == '1':
        opc_consulta = ''
        print('Indique la tarea que quiere consultar:  ')
        print('Todas las tareas:  1')
        print('En espera:  2')
        print('En ejecución:  3')
        print('Por aprobar:  4')
        print('Finalizada:  5')
        opc_consulta = input('Escriba la tarea que desea consultar: ')
        if opc_consulta == '1':
            print()
            print()
            print('** Consultando todas las tareas **')
            leer(Rut,'todo')

        elif opc_consulta == '2':
            print()
            print()
            print('** Consultando tareas en espera')
            leer(Rut, 'En espera')

        elif opc_consulta == '3':
            print()
            print()
            print('** Consultando tareas en ejecución')
            leer(Rut, 'En ejecución')

        elif opc_consulta == '4':
            print()
            print()
            print('** Consultando tareas por aprobar')
            leer(Rut, 'Por aprobar')

        elif opc_consulta == '5':
            print()
            print()
            print('** Consultando tareas finalizadas')
            leer(Rut, 'Finalizada')
        
    elif accion =='2':
        datosactualizados={'titulo':'','descripción':'','estado':'','fecha inicio':'','fecha finalizacion':''}
        print('** Actualizar Tarea **')
        print()
        id_Actualizar=int(input('Indique el id de la tarea que quiere actualizar:  '))
        print()
        print('** Nuevo Titulo **')
        print('**Nota: si no desea actalizar el titulo solo oprima ENTER**')
        datosactualizados['titulo'] = input('Indique el nuevo titulo de la tarea: ')
        print()
        print('** Nueva Descripción **')
        print('**Nota: si no desea actalizar la descripción solo oprima ENTER**')
        datosactualizados['descripcion'] = input('Indique la nueva descripción de la tarea: ')
        print()
        print('** Nuevo Estado **')
        print('En espera:  2')
        print('En ejecución:  3')
        print('Por aprobar:  4')
        print('Finalizada:  5')
        print('**Nota: si no desea actalizar el estado solo oprima ENTER**')
        estadonuevo=input('indique el nuevo estado de la tarea: ')
        if estadonuevo == '2':
            datosactualizados['estado']='En espera'
        elif estadonuevo == '3':
            datosactualizados['estado']='En ejecucion'
        elif estadonuevo == '4':
            datosactualizados['estado']='Por aprobar'
        elif estadonuevo == '5':
            now = datetime.now()
            datosactualizados['estado']='Finalizada'
            datosactualizados['fecha finalizacion']=str(now.day) + '/' + str(now.month) + '/' + str(now.year)
        
        now=datetime.now()
        datosactualizados['fecha inicio']=str(now.day) + '/' + str(now.month) + '/' + str(now.year)
        actualizar(Rut, id_Actualizar, datosactualizados)
        print()

    elif accion == '3':
        datosactualizados={'tarea':'','descripcion':'','estado':'','fecha inicio':'','fecha finalizacion':''}
        print('** Crear Nueva Tarea **')

        print()
        print('** titulo **')
        print()
        datosactualizados['titulo']=input('Indique el titulo de la tarea: ')
        print()
        print('** Descripción **')
        datosactualizados['descripcion']=input('Indique la descripción de la tarea:  ')
        print
        datosactualizados['estado']='En espera'
        now = datetime.now()
        datosactualizados['fecha inicio']=str(now.day) + '/' + str(now.month) + '/' + str(now.year)
        datosactualizados["fecha finalizacion"]=''
        agregrar(Rut,datosactualizados)
    elif accion=='4':
        print('')
        print('** Eliminar Tarea **')
        iden=int(input('Indique el ID de la tarea que desea eliminar: '))
        borrar(Rut,iden)