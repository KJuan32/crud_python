from datetime import datetime
from openpyxl import load_workbook


Rut="C:\\Users\\SENA\\Desktop\\git_taller\\crud_python\\supermercado_crud.xlsx"
Rut = r"C:\Users\SENA\Desktop\git_taller\crud_python\supermercado_crud.xlsx"
#direcciones de los archivos para poder ejecutar el codigo, importante

def leer(ruta:str, extraer:str):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['SuperM_de_crud']
    Hoja_datos = Hoja_datos['A2':'F' + str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:
        if isinstance(i[0].value,int):
            info.setdefault(i[0].value,{'nombre':i[1].value,  'categoria':i[2].value,
                                        'precio':i[3].value, 'cantidad':i[4].value,
                                        })
    
    if not(extraer == 'todo'):
        info = filtrar(info, extraer)
    
    for i in info:
        print('********Nombre********')
        print('Id: ' + str(i) + '\n' + 'Producto:  ' + str(info[i]['nombre']) + '\n' + 'Categoria: ' + str(info[i]['categoria']) + '\n'
              'Precio: ' + str(info[i]['precio']) + '\n' + 'Cantidad: ' + str(info[i]['cantidad']))    
        print()

    return


def filtrar(info:dict,filtro:str):
    aux={}

    for i in info:
        if info[i]['precio'] == filtro:
            aux.setdefault(i,info[i])

    return aux
    


def actualizar(ruta:str, identificador:int, datos_actualizados:dict):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['SuperM_de_crud']
    Hoja_datos = Hoja_datos['A2':'F' + str(Hoja_datos.max_row)]
    hoja = Archivo_Excel.active

    producto = 2
    categoria  = 3
    precio = 4
    cantidad = 5
    encontro = False
    for i in Hoja_datos:
        if i[0].value == identificador:
            fila=i[0].row   
            encontro = True
            for d in datos_actualizados:
                if d == 'producto' and not (datosactualizados[d] == ''):
                    hoja.cell(row = fila, column = producto).value= datosactualizados[d]
                elif d == 'categoria' and not (datosactualizados[d] == ''):
                    hoja.cell(row = fila, column = categoria).value= datosactualizados[d]
                elif d == 'precio' and not (datosactualizados[d] == ''):
                    hoja.cell(row = fila, column = precio).value= datosactualizados[d]
                elif d == 'cantidad' and not (datosactualizados[d] == ''):
                    hoja.cell(row = fila, column = cantidad).value= datosactualizados[d]
        

    Archivo_Excel.save(ruta)
    if encontro == False:
        print('Error: No existe una tarea con ese ID')
        print()
    return

def agregrar(ruta:int, datos:dict):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['SuperM_de_crud']
    Hoja_datos = Hoja_datos['A2':'F' + str(Hoja_datos.max_row+1)]
    hoja = Archivo_Excel.active

    producto = 2
    categoria = 3
    precio = 4
    cantidad = 5
    for i in Hoja_datos:

        if not( isinstance(i[0].value, int )) :
            identificador=i[0].row
            hoja.cell(row = identificador, column =1).value = identificador -1
            hoja.cell(row = identificador, column = producto).value = datos['producto']
            hoja.cell(row = identificador, column = categoria).value = datos['categoria']
            hoja.cell(row = identificador, column = precio).value = datos['precio']
            hoja.cell(row = identificador, column = cantidad).value = datos['cantidad']
            break
    Archivo_Excel.save(ruta)
    return

def borrar(ruta, identificador):
    Archivo_Excel = load_workbook(ruta)
    Hoja_datos = Archivo_Excel['SuperM_de_crud']
    Hoja_datos = Hoja_datos['A2':'F' + str(Hoja_datos.max_row)]
    hoja = Archivo_Excel.active

    producto = 2
    categoria = 3
    precio = 4
    cantidad = 5
    encontro = False
    for i in Hoja_datos:
        if i[0].value == identificador:
            fila = i[0].row
            encontro = True

            hoja.cell(row=fila, column = 1).value = ""
            hoja.cell(row=fila, column = producto).value = ""
            hoja.cell(row=fila, column = categoria).value = ""
            hoja.cell(row=fila, column = precio).value = ""
            hoja.cell(row=fila, column = cantidad).value = ""
    Archivo_Excel.save(ruta)
    if encontro == False:
        print("Error: No exite un producto con ese ID")
        print()
    return

Rut="C:\\Users\\SENA\\Desktop\\git_taller\\crud_python\\supermercado_crud.xlsx"
#direcciones de los archivos para poder ejecutar el codigo, importante



datosactualizados={'producto':'','categoria':'','precio':'','cantidad':''}
while True:
    print('Indique la acción que desea realizar: ')
    print('Consultar el producto: 1')
    print('Actualizar los productos: 2')
    print('Agregar nuevo producto:  3')
    print('Borrar productos:  4')
    accion = input('escriba la opción: ')

    if not(accion == '1') and not (accion == '2') and not (accion == '3') and not (accion == '4'):
        print('Comando invalido, por favor elija una opcion valida')
    elif accion == '1':
        opc_consulta = ''
        print('Indique eque quiere consultar:  ')
        print('Todos los productos:  1')
    
        opc_consulta = input('Escribar el producto que desea consultar: ')
        if opc_consulta == '1':
            print()
            print()
            print('** Consultando todos los productos **')
            leer(Rut,'todo')

        elif opc_consulta == '2':
            print()
            print()
            print('** Consultando la categoria')
            leer(Rut, 'Por categoria')

        elif opc_consulta == '3':
            print()
            print()
            print('** Consultando el precio')
            leer(Rut, 'Por precio')

        elif opc_consulta == '4':
            print()
            print()
            print('** Consultando la cantidad')
            leer(Rut, 'Por cantidad')

        
    elif accion =='2':
        datosactualizados={'producto':'','categoria':'','precio':'','cantidad':''}
        print('** Actualizar producto **')
        print()
        id_Actualizar=int(input('Indique el id del producto que quiere actualizar:  '))
        print()
        print('** El Nuevo nombre del producto **')
        print('**Nota: si no desea actalizar el producto solo oprima ENTER**')
        datosactualizados['producto'] = input('Indique el nuevo nombre del producto: ')
        print()
        print('** La Nueva categoria del producto **')
        print('**Nota: si no desea actalizar la descripción solo oprima ENTER**')
        datosactualizados['categoria'] = input('Indique la nueva categoria del producto: ')
        print()
        print('** El Nuevo precio del producto **')
        print('**Nota: si no desea actualizar el producto solo oprima ENTER**')
        datosactualizados['precio']=input('Indique el nuevo precio del producto: ')
        print()
        print('** La Nueva cantidad del producto**')
        print('**Nota: si no desea actualizar el producto solo oprima ENTER**')
        datosactualizados['cantidad']=input('indique la nueva cantidad del producto: ')
        print()
        
        actualizar(Rut, id_Actualizar, datosactualizados)
        print()

    elif accion == '3':
        datosactualizados={'producto':'','categoria':'','precio':'','cantidad':''}
        print('** Añadir nuevo producto **')

        print()
        print('** Nombre **')
        print()
        datosactualizados['producto']=input('Indique el nombre del producto: ')
        print()
        print('** Categoria **')
        datosactualizados['categoria']=input('Indique la categoria del producto:  ')
        print
        print('** Precio **')
        datosactualizados['precio']=input('Indique el precio del producto: ')
        print
        print('** Cantidad **')
        datosactualizados['cantidad']=input('Indique la cantidad del producto: ')
        agregrar(Rut,datosactualizados)
    elif accion=='4':
        print('')
        print('** Eliminar producto **')
        iden=int(input('Indique el ID del producto que desea eliminar: '))
        borrar(Rut,iden)